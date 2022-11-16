import openpyxl 

book=openpyxl.load_workbook("地形別面積.xlsx")
name=book.sheetnames
print(name)

hyo=book[name[0]]
print(hyo["A1"].value)
hyo["A1"]="都道府県名"
print(hyo["A1"].value)
print(hyo.max_row)
print(hyo.max_column)
book.save("地形別面積.xlsx")